"""
Builders for the GP, QC GP, SP, QC SP, and DD sheets in the generated fee
comparison workbook. Each mirrors its template counterpart's columns/headers,
but with two deliberate deviations (see conversation / build_fee_comparison.py
docstring for rationale):

  - Formulas that referenced external linked workbooks we don't have (QC GP,
    QC SP's CDCP/PT fee lookups) are replaced with plain literal values from
    our own extraction.
  - GP's Excel-Table structured-reference formulas ("Table1[[#This Row],...")
    are replaced with equivalent plain cell-reference formulas; GP's original
    last column contained an already-broken #REF! in the template itself,
    which is replaced with a self-consistent equivalent.

Claim-count inputs (claim weights, claim lines) are zeroed throughout, same
as the DH sheet -- that data isn't available from the CDCP/PT input files.
"""

import copy

from openpyxl.utils import get_column_letter


def copy_header_rows(ws, template_ws, num_header_rows: int, num_cols: int) -> None:
    for r in range(1, num_header_rows + 1):
        for c in range(1, num_cols + 1):
            src = template_ws.cell(r, c)
            dst = ws.cell(r, c, src.value)
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format
            dst.border = copy.copy(src.border)
    for merge_range in template_ws.merged_cells.ranges:
        if merge_range.max_row <= num_header_rows:
            ws.merge_cells(str(merge_range))
    for col, dim in template_ws.column_dimensions.items():
        ws.column_dimensions[col].width = dim.width
    # Freeze just below the header row(s) so it stays visible while
    # scrolling. Deliberately NOT copying template_ws.freeze_panes: the
    # template's freeze position is leftover scroll-state (e.g. DH freezes
    # everything above row 437, SP above row 15688) rather than a real
    # "keep the header visible" freeze, and blindly copying it onto a
    # differently-sized generated sheet is what caused sheets to appear
    # stuck / unable to scroll.
    ws.freeze_panes = f"A{num_header_rows + 1}"
    if template_ws.row_dimensions[1].height:
        ws.row_dimensions[1].height = template_ws.row_dimensions[1].height


def write_row(ws, row_num: int, values: list, style_row_cells: dict[int, object]) -> None:
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row_num, col_idx, value)
        style_src = style_row_cells[col_idx]
        cell.font = copy.copy(style_src.font)
        cell.fill = copy.copy(style_src.fill)
        cell.alignment = copy.copy(style_src.alignment)
        cell.number_format = style_src.number_format


def _style_row_cells(template_ws, style_row: int, num_cols: int) -> dict[int, object]:
    return {col: template_ws.cell(style_row, col) for col in range(1, num_cols + 1)}


NA = "N/A"


def dh_row_values(r: int, province: str, code: str, cdcp_2026, pt_2026) -> list:
    return [
        code,                                                 # A Procedure Code
        "DH",                                                 # B Specialty
        province,                                             # C PT
        f"=A{r}&C{r}&B{r}",                                   # D Helper
        NA,                                                   # E 2025 CDCP Fee
        cdcp_2026 if cdcp_2026 is not None else NA,           # F 2026 CDCP Fee
        NA,                                                   # G 2025 PT Fee
        pt_2026 if pt_2026 is not None else NA,               # H 2026 PT Fee
        f'=IFERROR((F{r}-E{r})/E{r},"N/A")',                  # I Unweighted Increase in CDCP Fees
        f'=IFERROR((H{r}-G{r})/G{r},"N/A")',                  # J Unweighted Increase in PT Fees
        f'=IFERROR(F{r}/H{r},"N/A")',                         # K Unweighted CDCP/PT 2026
        0,                                                     # L CDCP Claim Weight
        f'=IF(H{r}="N/A",0,L{r})',                            # M CDCP Claim Count for Share Covered
        f"=IFERROR(M{r}/VLOOKUP(C{r},'Claim Lines'!A$26:F$38,6,FALSE),\"\")",  # N PT Weight Share Covered
        0,                                                     # O Claim Count Nat'l
        0,                                                     # P PT Weight
        f"=IFERROR(P{r}*I{r},0)",                             # Q PT Weighted CDCP Increase
        f"=O{r}/'Claim Lines'!F$25",                          # R National Weight
        f"=IFERROR(R{r}*I{r},0)",                             # S National Weighted CDCP Increase
        f'=IFERROR(K{r}*R{r},"N/A")',                         # T National Weighted Share Covered
        f'=IFERROR(J{r}*R{r},"N/A")',                         # U PT Fee Increase Weighted
        f'=IFERROR(J{r}*P{r},"N/A")',                         # V PT Weighted PT Increase
        f"=L{r}/'Claim Lines'!H$3",                           # W All Provider Weight
        f"=IFERROR(W{r}*I{r},0)",                             # X All Provider Weighted Increase in CDCP Fees
        f'=IFERROR(N{r}*K{r},"")',                            # Y Share Covered
    ]


def build_dh_sheet(wb_new, template_ws, rows: list[tuple[str, str, float | None, float | None]]):
    """rows: list of (province, code, cdcp_2026, pt_2026)."""
    num_cols = 25
    ws = wb_new.create_sheet("DH")
    copy_header_rows(ws, template_ws, num_header_rows=1, num_cols=num_cols)
    style_cells = _style_row_cells(template_ws, style_row=2, num_cols=num_cols)

    r = 2
    for province, code, cdcp_2026, pt_2026 in rows:
        write_row(ws, r, dh_row_values(r, province, code, cdcp_2026, pt_2026), style_cells)
        r += 1


def gp_row_values(r: int, province: str, code: str, cdcp_2026, pt_2026) -> list:
    return [
        province,                                          # A PT
        "GP",                                               # B Specialty
        code,                                                # C Procedure Code
        f"=C{r}&A{r}",                                       # D Column2 (helper)
        f"=C{r}&A{r}&B{r}",                                  # E Column1 (helper)
        NA,                                                  # F 2025 CDCP Fee
        cdcp_2026 if cdcp_2026 is not None else NA,          # G 2026 CDCP Fee
        NA,                                                  # H 2025 PT Fee
        pt_2026 if pt_2026 is not None else NA,              # I 2026 PT Fee
        f'=IFERROR((G{r}-F{r})/F{r},"N/A")',                 # J Unweighted Increase in CDCP Fees
        f'=IFERROR(G{r}/I{r},"N/A")',                        # K Unweighted CDCP/PT 2026
        0,                                                    # L CDCP Claims for CDCP Fee Growth
        f'=IFERROR(L{r}/VLOOKUP(A{r},\'Claim Lines\'!A$2:B$16,2,FALSE),"")',  # M CDCP Claim Weight for CDCP Fee Growth
        f'=IFERROR(M{r}*J{r},"")',                           # N CDCP PT Weighted Fee Growth
        0,                                                    # O CDCP Claim Weight Nat'l
        f"=IFERROR(O{r}/'Claim Lines'!B$25,0)",              # P National Weight for Share Covered
        f'=IFERROR(P{r}*J{r},0)',                            # Q National Weighted CDCP Increase
        f"=O{r}/'Claim Lines'!H$3",                          # R All Providers Weight
        f'=IFERROR(R{r}*J{r},0)',                            # S All Providers CDCP Increase
        f'=IFERROR(P{r}*K{r},"N/A")',                        # T National Weighted Share Covered (TBC)
        f'=IFERROR(M{r}*K{r},"")',                           # U National Weight For PT Growth (was a hard #REF! in the template)
    ]


def build_gp_sheet(wb_new, template_ws, rows: list[tuple[str, str, float | None]]):
    """rows: list of (province, code, cdcp_2026, pt_2026) tuples, excluding QC."""
    num_cols = 21
    ws = wb_new.create_sheet("GP")
    copy_header_rows(ws, template_ws, num_header_rows=1, num_cols=num_cols)
    style_cells = _style_row_cells(template_ws, style_row=2, num_cols=num_cols)

    r = 2
    for province, code, cdcp_2026, pt_2026 in rows:
        write_row(ws, r, gp_row_values(r, province, code, cdcp_2026, pt_2026), style_cells)
        r += 1


def qc_gp_row_values(r: int, code: str, cdcp_2026, pt_2026) -> list:
    return [
        "QC",                                                # A PT
        "GP",                                                # B Specialty
        code,                                                # C Procedure Code
        f"=C{r}&A{r}&B{r}",                                  # D helper
        NA,                                                  # E 2025 CDCP Fee
        cdcp_2026 if cdcp_2026 is not None else NA,          # F 2026 CDCP Fee (was external-workbook VLOOKUP)
        NA,                                                  # G 2025 PT Fee (was external-workbook VLOOKUP)
        pt_2026 if pt_2026 is not None else NA,              # H 2026 PT Fee
        f'=IFERROR((F{r}-E{r})/E{r},"N/A")',                 # I Unweighted Increase in CDCP Fees
        f'=IFERROR(F{r}/H{r},"N/A")',                        # J Unweighted CDCP/PT 2026
        0,                                                    # K CDCP Claim Lines
        f"=K{r}/'Claim Lines'!B$12",                         # L CDCP Claim Weight for CDCP Fee Growth
        f'=IFERROR(L{r}*I{r},"")',                           # M PT Weighted CDCP Fee Growth
        f"=K{r}/'Claim Lines'!D$34",                         # N Claim Weight for GPSP Increase
        f'=IFERROR(N{r}*I{r},"")',                           # O PT Weighted CDCP Fee Growth of GPSP
        f"=IFERROR(K{r}/'Claim Lines'!H$3,0)",               # P National Weight
        f'=IFERROR(I{r}*P{r},0)',                            # Q National Weighted CDCP Increase
        f"=K{r}/'Claim Lines'!$H$3",                         # R All Providers Weight
        f'=IFERROR(R{r}*I{r},0)',                            # S All Providers CDCP Increase
        f'=IFERROR(P{r}*J{r},"N/A")',                        # T National Weighted Share Covered
        f'=IFERROR(L{r}*J{r},"")',                           # U PT Weighted Share Covered
        f"=IFERROR(K{r}/'Claim Lines'!D$34,\"\")",           # V Claim Weight for GPSP Share Covered
        f'=IFERROR(V{r}*J{r},"")',                           # W GPSP Share Covered
        None,                                                 # X (unused in template)
    ]


def build_qc_gp_sheet(wb_new, template_ws, rows: list[tuple[str, float | None]]):
    """rows: list of (code, cdcp_2026, pt_2026) tuples for QC only."""
    num_cols = 24
    ws = wb_new.create_sheet("QC GP")
    copy_header_rows(ws, template_ws, num_header_rows=1, num_cols=num_cols)
    style_cells = _style_row_cells(template_ws, style_row=2, num_cols=num_cols)

    r = 2
    for code, cdcp_2026, pt_2026 in rows:
        write_row(ws, r, qc_gp_row_values(r, code, cdcp_2026, pt_2026), style_cells)
        r += 1


def sp_row_values(r: int, province: str, sub_specialty: str, code: str, cdcp_2026, pt_2026) -> list:
    return [
        province,                                            # A Province
        sub_specialty,                                       # B Specialty
        code,                                                 # C Procedure Code
        f"=C{r}&A{r}",                                        # D helper
        f"=D{r}&B{r}",                                        # E Concat
        NA,                                                   # F 2025 CDCP Fee
        cdcp_2026 if cdcp_2026 is not None else NA,           # G 2026 CDCP Fee
        NA,                                                   # H 2025 PT Fee
        pt_2026 if pt_2026 is not None else NA,               # I 2026 PT Fee
        f'=IFERROR((G{r}-F{r})/F{r},"N/A")',                  # J Unweighted Increase in CDCP Fees
        f'=IFERROR(G{r}/I{r},"")',                            # K Unweighted CDCP/PT 2026
        0,                                                     # L CDCP Claim Weight
        0,                                                     # M CDCP Claim Weight (claim count)
        f"=IFERROR(M{r}/'Claim Lines'!C$25,0)",               # N National Weight
        f'=IFERROR(N{r}*J{r},0)',                             # O National Weighted CDCP Increase
        f"=L{r}/'Claim Lines'!$H$3",                          # P All Providers Weight
        f'=IFERROR(J{r}*P{r},0)',                             # Q All Providers CDCP Increase
        f'=IFERROR(N{r}*K{r},"N/A")',                         # R National Weighted Share Covered
    ]


def build_sp_sheet(wb_new, template_ws, rows: list[tuple[str, str, str, float | None]]):
    """rows: list of (province, sub_specialty, code, cdcp_2026, pt_2026), excluding QC."""
    num_cols = 18
    ws = wb_new.create_sheet("SP")
    copy_header_rows(ws, template_ws, num_header_rows=1, num_cols=num_cols)
    style_cells = _style_row_cells(template_ws, style_row=2, num_cols=num_cols)

    r = 2
    for province, sub_specialty, code, cdcp_2026, pt_2026 in rows:
        write_row(ws, r, sp_row_values(r, province, sub_specialty, code, cdcp_2026, pt_2026), style_cells)
        r += 1


def qc_sp_row_values(r: int, sub_specialty: str, code: str, cdcp_2026, pt_2026) -> list:
    return [
        "QC",                                                 # A Province
        sub_specialty,                                        # B Specialty
        code,                                                  # C Procedure Code
        f"=C{r}&A{r}",                                         # D helper
        f"=D{r}&B{r}",                                         # E helper2
        NA,                                                    # F 2025 CDCP Fee
        cdcp_2026 if cdcp_2026 is not None else NA,            # G 2026 CDCP Fee (was external-workbook VLOOKUP)
        NA,                                                    # H 2025 PT Fee
        pt_2026 if pt_2026 is not None else NA,                # I 2026 PT Fee (was external-workbook VLOOKUP)
        f'=IFERROR((G{r}-F{r})/F{r},"N/A")',                   # J Unweighted Increase in CDCP Fees
        f'=IFERROR(G{r}/I{r},"")',                             # K Unweighted CDCP/PT 2026
        0,                                                      # L CDCP Claims for CDCP Fee Growth
        f"=L{r}/'Claim Lines'!C$12",                           # M Claim Weight for CDCP Fee Growth
        f'=IFERROR(M{r}*J{r},"")',                             # N PT Weighted CDCP Fee Growth
        f"=L{r}",                                               # O CDCP Claims for Share Covered
        f'=IFERROR(M{r}*K{r},"")',                             # P PT Weighted Share Covered
        f"=L{r}/'Claim Lines'!D$12",                           # Q Claim Weight for CDCP Fee Growth GPSP
        f'=IFERROR(Q{r}*J{r},"")',                             # R PT Weighted CDCP Fee Growth GPSP
        f"=O{r}/'Claim Lines'!D$34",                           # S Claim Weight for Share Covered GPSP
        0,                                                      # T CDCP Claim Weight Nat'l
        f"=IFERROR(T{r}/'Claim Lines'!C$25,0)",                # U National Weight
        f'=IFERROR(U{r}*J{r},0)',                              # V National Weighted CDCP Increase
        f"=L{r}/'Claim Lines'!H$3",                            # W All Providers Weight
        f'=IFERROR(W{r}*J{r},0)',                              # X All Proivders CDCP Increase
        f'=IFERROR(U{r}*K{r},"")',                             # Y National Weighted Share Covered
        f'=IFERROR(S{r}*K{r},"")',                             # Z GPSP Share Covered
    ]


def build_qc_sp_sheet(wb_new, template_ws, rows: list[tuple[str, str, float | None]]):
    """rows: list of (sub_specialty, code, cdcp_2026, pt_2026) for QC only."""
    num_cols = 26
    ws = wb_new.create_sheet("QC SP")
    copy_header_rows(ws, template_ws, num_header_rows=1, num_cols=num_cols)
    style_cells = _style_row_cells(template_ws, style_row=2, num_cols=num_cols)

    r = 2
    for sub_specialty, code, cdcp_2026, pt_2026 in rows:
        write_row(ws, r, qc_sp_row_values(r, sub_specialty, code, cdcp_2026, pt_2026), style_cells)
        r += 1


def dd_row_values(r: int, province: str, code: str, cdcp_prof, cdcp_lab, pt_prof, pt_lab, pt_combo) -> list:
    cdcp_combo = None
    if cdcp_prof is not None:
        cdcp_combo = cdcp_prof + (cdcp_lab or 0)

    return [
        province,                                             # A PT
        "DD",                                                 # B Specialty
        code,                                                  # C Procedure Code
        f"=C{r}&A{r}&B{r}",                                    # D helper
        NA, NA, NA,                                            # E,F,G 2025 CDCP Prof/Lab/Combo
        cdcp_prof if cdcp_prof is not None else NA,            # H 2026 CDCP Prof Fee
        cdcp_lab if cdcp_lab is not None else NA,              # I 2026 CDCP Internal Lab Fee
        cdcp_combo if cdcp_combo is not None else NA,          # J 2026 CDCP Combo Fee
        NA, NA, NA,                                            # K,L,M 2025 PT Prof/Lab/Combo
        pt_prof if pt_prof is not None else NA,                # N 2026 PT Prof Fee
        pt_lab if pt_lab is not None else NA,                  # O 2026 PT Internal Lab Fee
        pt_combo if pt_combo is not None else NA,              # P 2026 PT Combo Fee
        f'=IFERROR((H{r}-E{r})/E{r},"N/A")',                   # Q Unweighted Increase CDCP Prof
        f'=IFERROR((I{r}-F{r})/F{r},"N/A")',                   # R Unweighted Increase CDCP Lab
        f'=IFERROR((J{r}-G{r})/G{r},"N/A")',                   # S Unweighted Increase CDCP Combo
        f'=IFERROR((N{r}-K{r})/K{r},"N/A")',                   # T Unweighted Increase PT Prof
        f'=IFERROR((O{r}-L{r})/L{r},"N/A")',                   # U Unweighted Increase PT Lab
        f'=IFERROR((P{r}-M{r})/M{r},"N/A")',                   # V Unweighted Increase PT Combo
        f'=IFERROR(J{r}/P{r},"N/A")',                          # W Unweighted CDCP/PT 2026 (Combo)
        0,                                                       # X CDCP Claim Weight (Claim Count)
        0,                                                       # Y Claim Count N/A Removed
        f"=IFERROR(X{r}/VLOOKUP(A{r},'Claim Lines'!A:E,5,FALSE),0)",  # Z PT Weight
        f"=Y{r}/'Claim Lines'!E$25",                           # AA National Weight
        f"=X{r}",                                               # AB National Weight PT
        f"=AB{r}/'Claim Lines'!E$25",                          # AC National Weight Share Covered
        f"=X{r}/'Claim Lines'!H$3",                            # AD All Provider Weight
        f'=IFERROR(Z{r}*Q{r},0)',                              # AE PT Weighted Increase CDCP Prof
        f'=IFERROR(Z{r}*R{r},0)',                              # AF Lab
        f'=IFERROR(Z{r}*S{r},0)',                              # AG Combo
        f'=IFERROR($AA{r}*Q{r},0)',                            # AH National Weighted Increase CDCP Prof
        f'=IFERROR($AA{r}*R{r},0)',                            # AI Lab
        f'=IFERROR($AA{r}*S{r},0)',                            # AJ Combo
        None, None,                                             # AK, AL (unused in template)
        f'=IFERROR($AA{r}*V{r},0)',                            # AM National Weighted Increase PT Combo
        f'=IFERROR(AD{r}*S{r},0)',                             # AN All Provider Weighted Increase Combo
        f'=IFERROR(W{r}*AC{r},"N/A")',                         # AO National Weighted Share Covered
        None, None,                                             # AP, AQ (unused in template)
    ]


def build_dd_sheet(wb_new, template_ws, rows: list[tuple[str, str, tuple, tuple]]):
    """rows: list of (province, code, (cdcp_prof, cdcp_lab), (pt_prof, pt_lab, pt_combo))."""
    num_cols = 43
    ws = wb_new.create_sheet("DD")
    copy_header_rows(ws, template_ws, num_header_rows=2, num_cols=num_cols)
    style_cells = _style_row_cells(template_ws, style_row=3, num_cols=num_cols)

    r = 3
    for province, code, (cdcp_prof, cdcp_lab), (pt_prof, pt_lab, pt_combo) in rows:
        write_row(ws, r, dd_row_values(r, province, code, cdcp_prof, cdcp_lab, pt_prof, pt_lab, pt_combo), style_cells)
        r += 1


def copy_claim_lines_sheet(wb_new, template_ws) -> None:
    """The Claim Lines sheet is static reference data (national/provincial
    claim volumes) plus formulas that reference our own DD/DH/GP/SP/QC sheets
    -- copy it verbatim so all those cross-references keep working."""
    ws = wb_new.create_sheet("Claim Lines")
    max_row = template_ws.max_row
    max_col = template_ws.max_column
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            src = template_ws.cell(r, c)
            if src.value is None:
                continue
            dst = ws.cell(r, c, src.value)
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format
    for col, dim in template_ws.column_dimensions.items():
        ws.column_dimensions[col].width = dim.width
