"""
Loaders for the CDCP price files, one per specialty "shape":

- GP and DH: the 'Specialty' column literally equals the specialty code
  ("GP", "DH"), so one code maps to one fee.
- SP: the 'Specialty' column holds a *sub-specialty* code (PA, EN, OS, ...)
  instead of the literal "SP" -- and the same procedure code commonly
  repeats under several different sub-specialties with different fees. So SP
  rows are returned as a flat list of (code, sub_specialty, fee) rather than
  a code -> fee dict.
- DD: has two separate fee columns (Provider Fee = professional fee, and
  Internal Lab Fee), which the output sheet keeps as separate Prof/Lab/Combo
  columns.
"""

from pathlib import Path

import openpyxl

from fee_extraction import normalize_code

# Yukon's CDCP price file is named/coded "YK" internally (both the filename
# and the 'Province' column value), while everywhere else in this project
# (province lists, PT fee guide folder names, the template's output labels)
# uses "YT". Without this, looking up "YT" finds no file at all.
CDCP_PROVINCE_ALIASES: dict[str, str] = {"YT": "YK"}


def _load_sheet(cdcp_dir: Path, province: str, sheet_name: str):
    file_province = CDCP_PROVINCE_ALIASES.get(province, province)
    path = cdcp_dir / f"2026 - CDCP PRICE FILE - {file_province}.xlsx"
    if not path.exists():
        return None
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}
    return ws, idx, file_province


def load_cdcp_simple_fees(
    cdcp_dir: Path, province: str, sheet_name: str, specialty: str, require_fee: bool = True
) -> dict[str, float | None]:
    """GP/DH-style: one 2026 fee per procedure code.

    `require_fee` controls what happens to codes with no Provider Fee in the
    CDCP file (e.g. an "I.C." / individually-costed service). DH's sheet in
    the template excludes such codes entirely; GP's sheet instead keeps them
    as a row with an "N/A" fee. Pass require_fee=False to match GP's
    convention.
    """
    result = _load_sheet(cdcp_dir, province, sheet_name)
    if result is None:
        return {}
    ws, idx, data_province = result

    fees: dict[str, float | None] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["Province"]] != data_province or row[idx["Specialty"]] != specialty:
            continue
        fee = row[idx["Provider Fee"]]
        if fee is None and require_fee:
            continue
        code = normalize_code(row[idx["Procedure Code"]])
        if code is None:
            continue
        fees[code] = float(fee) if fee is not None else None
    return fees


def load_cdcp_sp_rows(cdcp_dir: Path, province: str, require_fee: bool = True) -> list[tuple[str, str, float | None]]:
    """SP-style: every (code, sub-specialty) row for the province, since the
    same code can have a different fee under different sub-specialties.
    See load_cdcp_simple_fees for `require_fee`."""
    result = _load_sheet(cdcp_dir, province, "SP")
    if result is None:
        return []
    ws, idx, data_province = result

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["Province"]] != data_province:
            continue
        fee = row[idx["Provider Fee"]]
        if fee is None and require_fee:
            continue
        code = normalize_code(row[idx["Procedure Code"]])
        sub_specialty = row[idx["Specialty"]]
        if code is None or not sub_specialty:
            continue
        rows.append((code, sub_specialty, float(fee) if fee is not None else None))
    return rows


def load_cdcp_dd_fees(cdcp_dir: Path, province: str) -> dict[str, tuple[float, float]]:
    """DD-style: (professional fee, internal lab fee) per procedure code."""
    result = _load_sheet(cdcp_dir, province, "DD")
    if result is None:
        return {}
    ws, idx, data_province = result

    fees: dict[str, tuple[float, float]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["Province"]] != data_province or row[idx["Specialty"]] != "DD":
            continue
        prof_fee = row[idx["Provider Fee"]]
        if prof_fee is None:
            continue
        code = normalize_code(row[idx["Procedure Code"]])
        if code is None:
            continue
        lab_fee = row[idx["Internal Lab Fee"]] or 0
        fees[code] = (float(prof_fee), float(lab_fee))
    return fees
