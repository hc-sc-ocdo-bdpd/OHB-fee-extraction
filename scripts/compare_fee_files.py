"""
Compare a generated Fee Comparison workbook against the ground-truth workbook
and report a match rate for the '2026 CDCP Fee' and '2026 PT Fee' columns.

Sheets compared: GP, QC GP, SP, QC SP, DD, DH.
Rows are matched between the two files using (PT, Specialty, Procedure Code).

Blank/empty cells are treated as equivalent to the text 'N/A'. Numeric fees
are compared to the nearest cent to absorb float rounding noise, and a
'$'-formatted string value (e.g. "$89.80") is treated the same as the plain
number 89.8 -- a leading "$" (and thousands commas) don't create a mismatch.

Match rules:
  - If the ground truth has no real value (blank/N/A/0/0.0) for a field,
    the row is always counted as a match regardless of what the generated
    file has there -- there's nothing meaningful to disagree with.
  - 0/0.0 is treated as equivalent to 'N/A' on the ground-truth side for the
    rule above (a ground truth of 0 also counts as "no real value").
  - Rows present in the generated file but absent from the ground truth
    (extra keys) are counted and shown alongside the 'Missing' column, but
    never factored into the Total/Matched/Match % figures.

The 'DD' sheet has three sub-columns per fee ('Prof Fee', 'Internal Lab Fee',
'Combo Fee') under each of '2026 CDCP Fee' and '2026 PT Fee', and uses two
header rows instead of one -- these are each compared separately.

The 'DH' sheet's header labels for the first and third columns are swapped
relative to their actual data (column A holds the Procedure Code despite
being labelled 'PT', column C holds the PT despite being labelled
'Procedure Code') in both files. That layout is hardcoded below rather than
trusted from the header text.

Set GT_FILE and GEN_FILE below to the two workbook paths before running.
Also writes a full mismatch/missing/extra report to an .xlsx file next to
GT_FILE, in the same 'Output_2026 Fee Comparison' folder (see OUTPUT_FILE).
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font

GT_FILE = Path(r"C:\Users\JOGILL\OneDrive - HC-SC PHAC-ASPC\Desktop\OHB\Data\Output_2026 Fee Comparison\2026 Fee Comparisons v2_updated.xlsx")
GEN_FILE = Path(r"C:\Users\JOGILL\OneDrive - HC-SC PHAC-ASPC\Desktop\OHB\Data\Output_2026 Fee Comparison\2026 Fee Comparisons - generated.xlsx")

# Written into the same "Output_2026 Fee Comparison" folder as GT_FILE.
OUTPUT_FILE = GT_FILE.parent / "Fee_Comparison_Mismatches.xlsx"

SHEETS = ["GP", "QC GP", "SP", "QC SP", "DD", "DH"]

# Sheets whose identity columns (PT, Specialty, Procedure Code) are NOT in
# that order, as 1-indexed column numbers: (pt_col, specialty_col, code_col).
IDENTITY_COLUMN_OVERRIDES = {
    "DH": (3, 2, 1),
}


def normalize_code(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip().upper()


def normalize_fee(value):
    if value is None:
        return "N/A"
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "" or stripped.upper() == "N/A":
            return "N/A"
        # A currency-formatted value ("$89.80", "$1,234.56") is the same fee
        # as the plain number -- strip "$" and thousands commas and try to
        # read it as a number before giving up and treating it as text.
        cleaned = stripped.replace("$", "").replace(",", "").strip()
        try:
            return round(float(cleaned), 2)
        except ValueError:
            return stripped
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    return value


def is_blank_equivalent(value):
    """True for a ground-truth value that means 'no real fee' -- blank/N/A
    or a literal 0/0.0, both of which are treated the same way."""
    return value == "N/A" or (isinstance(value, (int, float)) and float(value) == 0.0)


def values_match(gt_value, gen_value):
    if gt_value == gen_value:
        return True
    # No real ground-truth value to disagree with -> always a match,
    # whatever the generated file has there.
    return is_blank_equivalent(gt_value)


def find_header_columns(sheet, header_row, target_names):
    """Return {name: 1-indexed column} for names found in the given header row."""
    found = {}
    for cell in sheet[header_row]:
        if cell.value in target_names:
            found[cell.value] = cell.column
    return found


def build_sheet_index(sheet, sheet_name):
    """
    Returns (rows_by_key, field_columns) where:
      rows_by_key: {(pt, specialty, code): {field_name: raw_cell_value}}
      field_columns: list of field names compared for this sheet
    """
    pt_col, spec_col, code_col = IDENTITY_COLUMN_OVERRIDES.get(sheet_name, (1, 2, 3))

    if sheet_name == "DD":
        # Two header rows: row 1 has the grouping label ('2026 CDCP Fee'),
        # row 2 has the sub-label ('Prof Fee' / 'Internal Lab Fee' / 'Combo Fee').
        group_row, sub_row, data_start = 1, 2, 3
        groups = find_header_columns(sheet, group_row, {"2026 CDCP Fee", "2026 PT Fee"})
        field_columns = {}
        for group_name, start_col in groups.items():
            for offset, sub_label in enumerate(["Prof Fee", "Internal Lab Fee", "Combo Fee"]):
                col = start_col + offset
                actual_sub = sheet.cell(row=sub_row, column=col).value
                if actual_sub != sub_label:
                    raise ValueError(
                        f"{sheet_name}: expected '{sub_label}' under '{group_name}' at "
                        f"column {col}, found {actual_sub!r}"
                    )
                field_columns[f"{group_name} - {sub_label}"] = col
    else:
        data_start = 2
        header_cols = find_header_columns(sheet, 1, {"2026 CDCP Fee", "2026 PT Fee"})
        field_columns = {name: col for name, col in header_cols.items()}

    rows_by_key = {}
    for row in sheet.iter_rows(min_row=data_start):
        pt = normalize_text(row[pt_col - 1].value)
        specialty = normalize_text(row[spec_col - 1].value)
        code = normalize_code(row[code_col - 1].value)
        if not pt and not specialty and not code:
            continue
        key = (pt, specialty, code)
        rows_by_key[key] = {
            field: row[col - 1].value for field, col in field_columns.items()
        }

    return rows_by_key, list(field_columns.keys())


def compare_sheet(gt_sheet, gen_sheet, sheet_name):
    gt_rows, fields = build_sheet_index(gt_sheet, sheet_name)
    gen_rows, _ = build_sheet_index(gen_sheet, sheet_name)
    extra_keys = sorted(set(gen_rows) - set(gt_rows))

    results = []
    all_mismatches = []
    all_missing = []
    for field in fields:
        matched = mismatched = missing = 0
        for key, gt_field_values in gt_rows.items():
            gt_value = normalize_fee(gt_field_values[field])
            if key not in gen_rows:
                missing += 1
                all_missing.append({"field": field, "key": key, "ground_truth": gt_value})
                continue
            gen_value = normalize_fee(gen_rows[key][field])
            if values_match(gt_value, gen_value):
                matched += 1
            else:
                mismatched += 1
                all_mismatches.append({
                    "field": field, "key": key,
                    "ground_truth": gt_value, "generated": gen_value,
                })
        total = len(gt_rows)
        match_rate = (matched / total * 100) if total else float("nan")
        results.append({
            "sheet": sheet_name,
            "field": field,
            "total": total,
            "matched": matched,
            "mismatched": mismatched,
            "missing": missing,
            "extra": len(extra_keys),
            "match_rate": match_rate,
        })

    return results, extra_keys, all_mismatches, all_missing


def format_missing(missing, extra):
    if extra:
        return f"{missing} (+{extra} extra)"
    return str(missing)


def write_report(all_results, mismatches_by_sheet, missing_by_sheet, extra_by_sheet, output_path):
    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    summary_ws = wb.active
    summary_ws.title = "Summary"
    headers = ["Sheet", "Field", "Total", "Matched", "Mismatched", "Missing", "Extra", "Match %"]
    summary_ws.append(headers)
    for cell in summary_ws[1]:
        cell.font = bold
    for r in all_results:
        summary_ws.append([
            r["sheet"], r["field"], r["total"], r["matched"],
            r["mismatched"], r["missing"], r["extra"], round(r["match_rate"], 2),
        ])
    total_all = sum(r["total"] for r in all_results)
    matched_all = sum(r["matched"] for r in all_results)
    overall_rate = round(matched_all / total_all * 100, 2) if total_all else None
    summary_ws.append([])
    summary_ws.append(["OVERALL", "", total_all, matched_all, total_all - matched_all, "", "", overall_rate])
    for cell in summary_ws[summary_ws.max_row]:
        cell.font = bold
    for col, width in zip("ABCDEFGH", [10, 28, 8, 9, 11, 16, 8, 9]):
        summary_ws.column_dimensions[col].width = width

    mismatches_ws = wb.create_sheet("Mismatches")
    mismatches_ws.append(["Sheet", "Field", "PT", "Specialty", "Procedure Code", "Ground Truth", "Generated"])
    for cell in mismatches_ws[1]:
        cell.font = bold
    for sheet_name, rows in mismatches_by_sheet.items():
        for row in rows:
            pt, specialty, code = row["key"]
            mismatches_ws.append([
                sheet_name, row["field"], pt, specialty, code,
                row["ground_truth"], row["generated"],
            ])
    for col, width in zip("ABCDEFG", [10, 28, 8, 10, 14, 14, 14]):
        mismatches_ws.column_dimensions[col].width = width

    missing_ws = wb.create_sheet("Missing")
    missing_ws.append(["Sheet", "Field", "PT", "Specialty", "Procedure Code", "Ground Truth"])
    for cell in missing_ws[1]:
        cell.font = bold
    for sheet_name, rows in missing_by_sheet.items():
        for row in rows:
            pt, specialty, code = row["key"]
            missing_ws.append([sheet_name, row["field"], pt, specialty, code, row["ground_truth"]])
    for col, width in zip("ABCDEF", [10, 28, 8, 10, 14, 14]):
        missing_ws.column_dimensions[col].width = width

    extra_ws = wb.create_sheet("Extra in Generated")
    extra_ws.append(["Sheet", "PT", "Specialty", "Procedure Code"])
    for cell in extra_ws[1]:
        cell.font = bold
    for sheet_name, keys in extra_by_sheet.items():
        for pt, specialty, code in keys:
            extra_ws.append([sheet_name, pt, specialty, code])
    for col, width in zip("ABCD", [10, 8, 10, 14]):
        extra_ws.column_dimensions[col].width = width

    wb.save(output_path)


def main():
    gt_wb = openpyxl.load_workbook(GT_FILE, data_only=True)
    gen_wb = openpyxl.load_workbook(GEN_FILE, data_only=True)

    all_results = []
    mismatches_by_sheet = {}
    missing_by_sheet = {}
    extra_by_sheet = {}

    header = f"{'Sheet':<8} {'Field':<28} {'Total':>7} {'Matched':>8} {'Mismatch':>9} {'Missing':>16} {'Match %':>8}"
    print(header)
    print("-" * len(header))

    for sheet_name in SHEETS:
        if sheet_name not in gt_wb.sheetnames or sheet_name not in gen_wb.sheetnames:
            print(f"{sheet_name}: missing from one of the workbooks, skipped")
            continue
        results, extra_keys, mismatches, missing = compare_sheet(
            gt_wb[sheet_name], gen_wb[sheet_name], sheet_name
        )
        all_results.extend(results)
        mismatches_by_sheet[sheet_name] = mismatches
        missing_by_sheet[sheet_name] = missing
        extra_by_sheet[sheet_name] = extra_keys
        for r in results:
            missing_display = format_missing(r["missing"], r["extra"])
            print(f"{r['sheet']:<8} {r['field']:<28} {r['total']:>7} {r['matched']:>8} "
                  f"{r['mismatched']:>9} {missing_display:>16} {r['match_rate']:>7.2f}%")

    print("-" * len(header))
    total_all = sum(r["total"] for r in all_results)
    matched_all = sum(r["matched"] for r in all_results)
    overall_rate = (matched_all / total_all * 100) if total_all else float("nan")
    print(f"{'OVERALL':<8} {'':<28} {total_all:>7} {matched_all:>8} "
          f"{total_all - matched_all:>9} {'':>16} {overall_rate:>7.2f}%")

    write_report(all_results, mismatches_by_sheet, missing_by_sheet, extra_by_sheet, OUTPUT_FILE)
    print(f"\nFull mismatch/missing/extra report written to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()