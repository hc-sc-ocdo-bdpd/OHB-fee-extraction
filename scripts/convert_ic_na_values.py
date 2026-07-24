"""
Convert every literal 'I.C.'/'I.C' and '#N/A' cell value in an xlsx workbook
to 'N/A'.

Only touches cells whose value IS one of those strings (after trimming
whitespace and ignoring case) -- formulas are left untouched, including
IFERROR(...,"N/A") formulas that already evaluate to "N/A".

Some sheets (GP, QC GP) hold VLOOKUP formulas against an external linked
workbook that isn't present alongside this file (e.g.
"=VLOOKUP(E1108,[1]SK!$E:$I,5,FALSE)"). openpyxl drops the cached result of
every formula when it re-saves a workbook, and without the source file those
would recalculate to #NAME? and the link would be gone for good. To avoid
that, any such formula is frozen to its last cached value before saving.

Set INPUT_FILE below to the workbook's path. The output is written next to
it in the same folder, with "_updated" appended to the filename.
"""

import re
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "Data"

INPUT_FILE = DATA_DIR / "Output_2026 Fee Comparison" / "2026 Fee Comparisons v2.xlsx"

IC_VALUES = {"I.C.", "I.C"}
NA_VALUE = "#N/A"
EXTERNAL_LINK_RE = re.compile(r"^=.*\[\d+\]")


def convert_cell_values(workbook: openpyxl.Workbook) -> int:
    changed = 0
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                stripped = value.strip()
                if stripped.upper() in IC_VALUES or stripped == NA_VALUE:
                    cell.value = "N/A"
                    changed += 1
    return changed


def freeze_external_link_formulas(workbook: openpyxl.Workbook, cached_workbook: openpyxl.Workbook) -> int:
    frozen = 0
    for sheet in workbook.worksheets:
        cached_sheet = cached_workbook[sheet.title]
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and EXTERNAL_LINK_RE.match(value):
                    cell.value = cached_sheet[cell.coordinate].value
                    frozen += 1
    return frozen


def main():
    input_path = INPUT_FILE
    output_path = input_path.with_name(f"{input_path.stem}_updated{input_path.suffix}")

    workbook = openpyxl.load_workbook(input_path)
    cached_workbook = openpyxl.load_workbook(input_path, data_only=True)

    frozen = freeze_external_link_formulas(workbook, cached_workbook)
    changed = convert_cell_values(workbook)

    workbook.save(output_path)
    print(f"Converted {changed} cell(s) to 'N/A'.")
    if frozen:
        print(f"Froze {frozen} external-link formula(s) to their last cached value.")
    print(f"Saved to {output_path}")


if __name__ == "__main__":
    main()